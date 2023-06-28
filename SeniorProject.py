#Takes in player name, year, and first three scores of a tournament then predicts the last day with different ML models
import requests
import time
import json
import csv
import os
import pandas as pd
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor, GradientBoostingClassifier, GradientBoostingRegressor
from sklearn.svm import SVC, SVR
from lightgbm import LGBMClassifier, LGBMRegressor

#TODO:: Clean names

class player:

    def __init__(self):
        self.name = None
        self.score = None
        self.r1 = None
        self.r2 = None
        self.r3 = None
        self.r4 = None
        self.id = None

    #Print out obj
    def __str__(self):
        return f"player(name={self.name} )"

    #Getter methods
    def get_name(self):
        return self.name
    def get_score(self):
        return self.score
    def get_r1(self):
        return self.r1
    def get_r2(self):
        return self.r2
    def get_r3(self):
        return self.r3
    def get_r4(self):
        return self.r4
    def get_id(self):
        return self.id

    #Setter methods
    def name(self, value):
        self.name = value
    def score(self, value):
        self.score = value
    def weight(self, value):
        self.weight = value
    def r1(self, value):
        self.r1 = value
    def r2(self, value):
        self.r2 = value
    def r3(self, value):
        self.r3 = value
    def r4(self, value):
        self.r4 = value
    def id(self, value):
        self.id = value

class tournament_data:
    def __init__(self):
        self.name = None
        self.id = None
        self.day1 = None
        self.day2 = None
        self.day3 = None
        self.day4 = None
        self.par = None
        self.yardage = None
        self.rating = None
        self.written = None

    #Print out obj
    def __str__(self):
        return f"tournament_data(name={self.name}, id={self.id}), day1={self.day1}, day2={self.day2}, day3={self.day3}, day4={self.day4}"

    #Getter methods
    def get_name(self):
        return self.name
    def get_id(self):
        return self.id
    def get_day1(self):
        return self.day1
    def get_day2(self):
        return self.day2
    def get_day3(self):
        return self.day3
    def get_day4(self):
        return self.day4
    def get_par(self):
        return self.par
    def get_yardage(self):
        return self.yardage
    def get_rating(self):
        return self.rating

    #Setter methods
    def name(self, value):
        self.name = value
    def id(self, value):
        self.id = value
    def day1(self, value):
        self.day1 = value
    def day2(self, value):
        self.day2 = value
    def day3(self, value):
        self.day3 = value
    def day4(self, value):
        self.day4 = value
    def par(self, value):
        self.par = value
    def yardage(self, value):
        self.yardage = value
    def rating(self, value):
        self.rating = value

##
# Function to predict the last day of the tournament for a player
##
def predictions(filename, day1, day2, day3, day4, fullname, tournament_name):
    # Load the data containing the first three rounds and the corresponding outcomes
    data = pd.read_csv(filename)  # Replace file with your actual data file
    tournament_data = pd.read_csv("CourseData.csv")

    # Filter rows based on a condition (e.g., tournament_name equals "Tournament A")
    filtered_rows = tournament_data[tournament_data['Tournament Name'] == tournament_name]
    tourney_yardage = filtered_rows["Yardage"].values
    tourney_par = filtered_rows["Par"].values
    tourney_rating = filtered_rows["Rating"].values

    yarage_value = tourney_yardage[0]
    par_value = tourney_par[0]
    rating_value = tourney_rating[0]

    # Split the data into features (first three rounds) and target variable (fourth game outcome)
    #=========Day1==========
    #X = data[['Yardage', 'Par', 'Rating']]
    #y = data['day1']
    #new_game = [[yarage_value, par_value, rating_value]]

    #==========Day2==========
    #X = data[['day1', 'Yardage', 'Par', 'Rating']]
    #y = data['day2']
    #new_game = [[day1, yarage_value, par_value, rating_value]]

    #==========Day3==========
    #X = data[['day1', 'day2', 'Yardage', 'Par', 'Rating']]
    #y = data['day3']
    #new_game = [[day1, day2, yarage_value, par_value, rating_value]]

    #==========Day4==========
    X = data[['day1', 'day2', 'day3', 'Yardage', 'Par', 'Rating']]
    y = data['day4']
    new_game = [[day1, day2, day3, yarage_value, par_value, rating_value]]
    

    # Split the data into training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=.1, random_state=42)
    

    #LGBMClassifier, LGBMRegressor
    LGBMClassifier_model = LGBMClassifier()
    LGBMRegressor_model = LGBMRegressor()
    LGBMClassifier_model.fit(X_train.values, y_train)
    LGBMRegressor_model.fit(X_train.values, y_train)
    LGBMClassifier_model_prediction = LGBMClassifier_model.predict(new_game)
    LGBMRegressor_model_prediction = LGBMRegressor_model.predict(new_game)

    #SVC, SVR
    svc_model = SVC()
    svr_model = SVR()
    svc_model.fit(X_train.values, y_train)
    svr_model.fit(X_train.values, y_train)
    svc_model_prediction = svc_model.predict(new_game)
    svr_model_prediction = svr_model.predict(new_game)

    #RandomForestClassifier, RandomForestRegressor, GradientBoostingClassifier, GradientBoostingRegressor
    random_forest_classifier_model = RandomForestClassifier()
    random_forest_regressor_model = RandomForestRegressor()
    gradien_boosting_classifier_model = GradientBoostingClassifier()
    gradient_boosting_regressor_model = GradientBoostingRegressor()
    random_forest_classifier_model.fit(X_train.values, y_train)
    random_forest_regressor_model.fit(X_train.values, y_train)
    gradien_boosting_classifier_model.fit(X_train.values, y_train)
    gradient_boosting_regressor_model.fit(X_train.values, y_train)
    random_forest_classifier_model_prediction = random_forest_classifier_model.predict(new_game)
    random_forest_regressor_model_prediction = random_forest_regressor_model.predict(new_game)
    gradien_boosting_classifier_model_prediction = gradien_boosting_classifier_model.predict(new_game)
    gradient_boosting_regressor_model_prediction = gradient_boosting_regressor_model.predict(new_game)

    #Linear
    linear_regression_model = LinearRegression()
    linear_regression_model.fit(X_train.values, y_train)
    linear_regression_model_prediction = linear_regression_model.predict(new_game)

    #Averages
    linear_average = (linear_regression_model_prediction[0])
    random_forest_average = (random_forest_classifier_model_prediction[0] + random_forest_regressor_model_prediction[0]) / 2
    gradient_boosting_average = (gradien_boosting_classifier_model_prediction[0] + gradient_boosting_regressor_model_prediction[0]) / 2
    SVM_average = (svc_model_prediction[0] + svr_model_prediction[0]) / 2
    LGBM_average = (LGBMClassifier_model_prediction[0] + LGBMRegressor_model_prediction[0]) / 2

    average = (linear_average + random_forest_average + gradient_boosting_average + LGBM_average + SVM_average) / 5

    filename = "fourth.xlsx"
    fieldnames = ["Name", "Lin Regression", "Random Forest Class", "Random Forest Regress", "Gradient Boost Class",
                "Gradient Boost Regress", "SVC", "SVR", "LGBMClass", "LGBMRegress", "Linear Ave",
                "Random Forest Ave", "Gradient Boost Ave", "SVM Ave", "LGBM Ave", "Total Ave", "R1", "R2", "R3", "R4"]

    # Check if the file exists
    file_exists = os.path.isfile(filename)

    # Create a new workbook or load an existing one
    workbook = openpyxl.load_workbook(filename) if file_exists else openpyxl.Workbook()

    # Select the active sheet
    sheet = workbook.active

    # Write the header row if the file is empty
    if not file_exists:
        for col_num, fieldname in enumerate(fieldnames, start=1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = fieldname

    # Find the next available row
    next_row = sheet.max_row + 1

    # Write the data row
    data_row = [fullname, linear_regression_model_prediction[0], random_forest_classifier_model_prediction[0],
                random_forest_regressor_model_prediction[0], gradien_boosting_classifier_model_prediction[0],
                gradient_boosting_regressor_model_prediction[0], svc_model_prediction[0], svr_model_prediction[0],
                LGBMClassifier_model_prediction[0], LGBMRegressor_model_prediction[0], linear_average,
                random_forest_average, gradient_boosting_average, SVM_average, LGBM_average, average, day1, day2, day3, day4]

    for col_num, value in enumerate(data_row, start=1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}{next_row}"] = value

    # Save the workbook
    workbook.save(filename)
    print("Excel file created successfully.")


    #return svr_model_prediction[0]

    #MLP does different values every time
    """
    print("Linear Regression Model = ", str(linear_regression_model_prediction[0]))
    print("Random Forest Classifier Model = ", str(random_forest_classifier_model_prediction[0]), "\nRandom Forest Regressor Model = ", str(random_forest_regressor_model_prediction[0]))
    print("Gradien Boosting Classifier Model = ", str(gradien_boosting_classifier_model_prediction[0]), "\nGradient Boosting Regressor Model = ", str(gradient_boosting_regressor_model_prediction[0]))
    print("SVC Model = ", str(svc_model_prediction[0]), "\nSVR Model = ", str(svr_model_prediction[0]))
    print("LGBMClassifier Model = ", str(LGBMClassifier_model_prediction[0]), "\nLGBMRegressor Model = ", str(LGBMRegressor_model_prediction[0]))

    print("Linear Average = " , linear_average)
    print("Random Forest Average = ", random_forest_average)
    print("Gradient Boosting Avergae = ", gradient_boosting_average)
    print("SVM Average = ", SVM_average)
    

    print("Total Average = " , average)
    """
    print("LGBM Average = ", LGBM_average)

##
# Functions that gets a players past tournament data and puts it into a csv for easy reading
##
def data_retrieval(target_player_full_name, target_year, day1, day2, day3, day4, tournament_name):

    names = target_player_full_name.split()  # Split the full name by whitespace
    
    #TODO::Might not work with people with three names eg. Si Woo Kim
    first_name_initial = names[0][0]  # Extract the first letter of the first name
    last_name = names[-1] 
    csv_file = "CSV Folder/" + first_name_initial + "." + last_name + ".csv"
    course_data = pd.read_csv("CourseData.csv")
    
    #Already existing file for player
    if os.path.exists(csv_file):
        print("Loading Model...")
        return predictions(csv_file, day1, day2, day3, day4, target_player_full_name, tournament_name)
    
    #No existing model
    else:
        #API for player ids
        print("Loading Model...")

        req_id = requests.get('http://api.sportradar.us/golf/trial/pga/v3/en/2023/players/profiles.json?api_key=b8rb3fgtwdj76yd7vkgr4uy4')
        time.sleep(1) # Wait for 1 second (can't do multiple api calls at once)

        player_obj = player()

        # Successful request
        if(req_id.status_code == 200): 
            id_data = req_id.json()

            #Get Player id
            for i in id_data["players"]:
                if(target_player_full_name == i["first_name"] + " " + i["last_name"]):
                    player_obj.name = i["abbr_name"]
                    player_obj.id = i["id"]
                    
            print(player_obj.name)
        #Player not found
        if player_obj.name is None:
            error = "Player '" + target_player_full_name + "' Not Found"
            print(error)
            return error

        else:

            #API call to get all data for a player
            url = f'http://api.sportradar.us/golf/trial/v3/en/players/{player_obj.id}/profile.json?api_key=b8rb3fgtwdj76yd7vkgr4uy4'
            player_data = requests.get(url)
            if(player_data.status_code == 200): # Successful request
                player_data = player_data.json()
        
                #Get tournaments only in the correct year for specific player
                tournament_array = []
        
                for tournament in player_data["previous_tournaments"]:
                    seasons = tournament["seasons"]
                    for season in seasons: 
                        if(season["year"] == int(target_year)):
                                        
                            #Check if Cut PLyer is cut 
                            #position needs to be first or errors in reading
                            """
                            if(tournament["seasons"]["tour"]["alias"] == "LIV"):
                                tournament_obj = tournament_data()
                                tournament_obj.name = tournament["name"]

                                if(tournament_obj.name in course_data["Tournament Name"].values):
                                    print(tournament_obj.name)
                                   
                                    filtered_rows = course_data[course_data['Tournament Name'] == tournament_obj.name]

                                    # Iterate over the filtered rows and assign values to variables
                                    for index, row in filtered_rows.iterrows():
                                        tournament_obj.yardage = row['Yardage']
                                        tournament_obj.par = row['Par']
                                        tournament_obj.rating = row['Rating']
                                

                                    tournament_obj.id = tournament["id"]

                                    #Get strokes and stuff
                                    for i in range(0, 3):
                                        strokes_data = tournament["leaderboard"]["rounds"][i]
                                        setattr(tournament_obj, f"day{i+1}", strokes_data["strokes"])

                                    tournament_array.append(tournament_obj)
                            """
                            if("position" not in tournament["leaderboard"] or tournament["leaderboard"]["position"] > 72 or len(tournament["leaderboard"]["rounds"]) < 4):
                                # PLayer was cut, Do not use these data points
                                continue
                            else:  

                                #Get Tournament data 
                                tournament_obj = tournament_data()
                                tournament_obj.name = tournament["name"]

                                if(tournament_obj.name in course_data["Tournament Name"].values):
                                   
                                    filtered_rows = course_data[course_data['Tournament Name'] == tournament_obj.name]
                                    tournament_obj.id = tournament["id"]

                                    # Iterate over the filtered rows and assign values to variables
                                    for index, row in filtered_rows.iterrows():
                                        tournament_obj.yardage = row['Yardage']
                                        tournament_obj.par = row['Par']
                                        tournament_obj.rating = row['Rating']

                                    #Get strokes and stuff
                                    for i in range(0, 4):
                                        strokes_data = tournament["leaderboard"]["rounds"][i]
                                        setattr(tournament_obj, f"day{i+1}", strokes_data["strokes"])

                                    tournament_array.append(tournament_obj)

                if(len(tournament_array) < 3):
                    error = "Not enough Data"
                    print(error)
                    return error

                #Put data into csv file
                filename = "CSV Folder/" + player_obj.name + ".csv"
                fieldnames = ["Tournament Name", "day1", "day2", "day3", "day4", "Yardage", "Par", "Rating"]
                already_written = {}
                with open(filename, mode="w", newline="") as file:
                    writer = csv.DictWriter(file, fieldnames=fieldnames)
                    writer.writeheader()

                    for data in tournament_array:
                
                        if data.get_id() not in already_written.keys():
                            writer.writerow({
                                "Tournament Name": data.get_name(),
                                "day1": data.get_day1(),
                                "day2": data.get_day2(),
                                "day3": data.get_day3(),
                                "day4": data.get_day4(),
                                "Yardage": data.get_yardage(),
                                "Par": data.get_par(),
                                "Rating": data.get_rating(),
                            })
                            already_written[data.get_id()] = True
                    
                print("CSV file created successfully.")


            return predictions(filename, day1, day2, day3, day4, target_player_full_name, tournament_name)

def scrape_data():
    #Get leaderboad through espn
    url = 'https://www.espn.com/golf/leaderboard'
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    dfs = pd.read_html(url, header = 0)
    data = dfs[-1] #this is leaderboard


    #loop through and add player to array
    players = []
    for index, row in data.iterrows():

        """
        #Get everyone including after cut (Good for first 2 days)
        if(row['PLAYER'].split()[0] != "Projected"):
            player_data = player()
            player_data.name = row['PLAYER']
            player_data.score = row['SCORE']
            player_data.r1 = row['R1']
            player_data.r2 = row['R2']
            player_data.r3 = row['R3']
            player_data.r4 = row['R4']

            players.append(player_data)
        """
        
        #Get Everyone Before the cut then stops (Good for last 2 days)
        player_data = player()
        player_data.name = row['PLAYER']
        player_data.score = row['SCORE']
        player_data.r1 = row['R1']
        player_data.r2 = row['R2']
        player_data.r3 = row['R3']
        player_data.r4 = row['R4']
        players.append(player_data)

        if(row['PLAYER'].split()[0] == "The"):
            break
        
    
    return players

def player_csv_already_created(target_player_full_name, day1, day2, day3, day4, tournament_name):
    names = target_player_full_name.split()  # Split the full name by whitespace
    
    #TODO::Might not work with people with three names eg. Si Woo Kim
    first_name_initial = names[0][0]  # Extract the first letter of the first name
    last_name = names[-1] 
    csv_file = "CSV Folder/" + first_name_initial + "." + last_name + ".csv"
    course_data = pd.read_csv("CourseData.csv")
    
    #Already existing file for player
    if os.path.exists(csv_file):
        print("Loading Model...")
        return predictions(csv_file, day1, day2, day3, day4, target_player_full_name, tournament_name)

def add_scores_to_xlsx(target_player_full_name, day1, day2, day3, day4):
    filename = "deltas.xlsx"

    fieldnames = ["Name", "Lin Regression", "Random Forest Class", "Random Forest Regress", "Gradient Boost Class", "Gradient Boost Regress",
              "SVC", "SVR", "LGBMClass", "LGBMRegress", "Linear Ave", "Random Forest Ave", "Gradient Boost Ave", "SVM Ave", 
              "LGBM Ave", "Total Ave", "R1", "R2", "R3", "R4"]

    # Load the workbook
    workbook = load_workbook(filename)
    worksheet = workbook.worksheets[4]
    #print(worksheet.title)

    # Find the target player's row in the worksheet
    target_player_row_index = None
    for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == target_player_full_name:
            target_player_row_index = i
            break

    # Check if the target player's row is found
    if target_player_row_index is not None:
        # Convert the string numbers to the desired numeric type

        try:
            day1 = int(day1)
        except (ValueError, TypeError):
            day1 = 0

        try:
            day2 = int(day2)
        except (ValueError, TypeError):
            day2 = 0

        try:
            day3 = int(day3)
        except (ValueError, TypeError):
            day3 = 0

        try:
            day4 = int(day4)
        except (ValueError, TypeError):
            day4 = 0

        # Update the target player's scores for the new rounds
        worksheet.cell(row=target_player_row_index, column=fieldnames.index("R1") + 1, value=day1)  # R1 column
        worksheet.cell(row=target_player_row_index, column=fieldnames.index("R2") + 1, value=day2)  # R2 column
        worksheet.cell(row=target_player_row_index, column=fieldnames.index("R3") + 1, value=day3)  # R3 column
        worksheet.cell(row=target_player_row_index, column=fieldnames.index("R4") + 1, value=day4)  # R4 column

        # Save the workbook
        workbook.save(filename)


# Define the main function
def main():
    # Prompt the user for input
    
    
    players = scrape_data()
    tournament_name = "Travelers Championship"
    #data_retrieval("Rory McIlroy", 2023, 0, 0, 0, 0, tournament_name)

    for player in players:
        print(player.name)
        #print(player.name, player.r1, player.r2, player.r3, player.r4)
        #data_retrieval(player.name, 2023, player.r1, player.r2, player.r3, player.r4, tournament_name)
        #player_csv_already_created(player.name, int(player.r1), int(player.r2), int(player.r3), player.r4, tournament_name)
        add_scores_to_xlsx(player.name, player.r1, player.r2, player.r3, player.r4)
    

    """
    target_player_full_name = "Phil Mickelson"
    target_year = 2023
    tournament_name = "U.S. Open"
    day1 = 70
    day2 = 70
    day3 = 67
    data_retrieval(target_player_full_name, target_year, day1, day2, day3, tournament_name)
    """

    """    
    filename = "ConditionalTest.xlsx"

    # Load the workbook
    workbook = load_workbook(filename)

    # Select the desired worksheet
    worksheet = workbook.active

    # Red Fill
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")

    # Green Fill
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")

    # Yellow Fill
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_font = Font(color="9C6500")

    differential_style = DifferentialStyle(fill=green_fill, font=green_font)

    # Create a rule with the condition and differential style
    rule = Rule(type="cellIs", operator="lessThan", formula=["5"], dxf=differential_style)
    rule.formula = ["$D2<5"]  # Assuming "x+y" column is column D
    worksheet.conditional_formatting.add(f"D2:D{worksheet.max_row}", rule)

    # Save the workbook
    workbook.save(filename)
    """
    
    
# Call the main function to start the program
if __name__ == "__main__":
    main()
